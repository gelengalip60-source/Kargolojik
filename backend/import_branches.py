#!/usr/bin/env python3
"""
Script to import branch data from Excel files into MongoDB
"""

import asyncio
import os
import sys
import uuid
from datetime import datetime
import httpx
from openpyxl import load_workbook
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Excel file URLs
EXCEL_FILES = {
    'Aras Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/vf7twzgp_aras.xlsx',
    'PTT Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/xwmutxuy_ptt.xlsx',
    'Sürat Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/97ie2tcw_s%C3%BCrat.xlsx',
    'DHL Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/ryngoqij_dhl.xlsx',
    'Inter Global Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/arrfhpw9_inter.xlsx',
    'TNT Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/sev87zyu_tnt.xlsx',
    'UPS Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/fgvucihf_ups.xlsx',
    'Yurtiçi Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/5bvuzryg_yurti%C3%A7i.xlsx',
}

# Company logo URLs
COMPANY_LOGOS = {
    'Aras Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/k3zkcxdq_aras-kargo-logo-png_seeklogo-510325.png',
    'DHL Kargo': 'https://customer-assets.emergentagent.com/job_shipntracker/artifacts/eosc97km_dhl-logo-png_seeklogo-40800.png',
    'PTT Kargo': 'https://upload.wikimedia.org/wikipedia/commons/thumb/6/60/PTT_logo.svg/512px-PTT_logo.svg.png',
    'Sürat Kargo': 'https://upload.wikimedia.org/wikipedia/commons/thumb/c/c9/S%C3%BCrat_Kargo_logo.svg/512px-S%C3%BCrat_Kargo_logo.svg.png',
    'Yurtiçi Kargo': 'https://upload.wikimedia.org/wikipedia/commons/thumb/0/0b/Yurtici_Kargo_logo.svg/512px-Yurtici_Kargo_logo.svg.png',
    'UPS Kargo': 'https://upload.wikimedia.org/wikipedia/commons/thumb/6/6b/United_Parcel_Service_logo_2014.svg/512px-United_Parcel_Service_logo_2014.svg.png',
    'TNT Kargo': 'https://upload.wikimedia.org/wikipedia/commons/thumb/9/90/TNT_Express_Logo.svg/512px-TNT_Express_Logo.svg.png',
    'Inter Global Kargo': '',
}

async def download_file(url: str, filename: str) -> str:
    """Download a file from URL"""
    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.get(url)
        response.raise_for_status()
        
        filepath = f"/tmp/{filename}"
        with open(filepath, 'wb') as f:
            f.write(response.content)
        return filepath

def extract_branches_from_excel(filepath: str, company: str) -> list:
    """Extract branch data from Excel file"""
    branches = []
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Get header row to understand column mapping
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).lower() if cell.value else '')
    
    print(f"Headers found for {company}: {headers}")
    
    # Try to map columns
    name_col = None
    city_col = None
    district_col = None
    address_col = None
    phone_col = None
    
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        if 'sube_adi' in h_lower or 'şube' in h_lower or 'name' in h_lower:
            name_col = i
        elif 'sehir' in h_lower or 'şehir' in h_lower or 'city' in h_lower or h_lower == 'il':
            city_col = i
        elif 'ilce' in h_lower or 'ilçe' in h_lower or 'district' in h_lower:
            district_col = i
        elif 'adres' in h_lower or 'address' in h_lower:
            address_col = i
        elif 'telefon_1' in h_lower or 'telefon' in h_lower or 'phone' in h_lower:
            phone_col = i
    
    print(f"Column mapping: name={name_col}, city={city_col}, district={district_col}, address={address_col}, phone={phone_col}")
    
    # Read data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            # Get values with defaults
            name = str(row[name_col].value) if name_col is not None and row[name_col].value else ''
            city = str(row[city_col].value) if city_col is not None and row[city_col].value else ''
            district = str(row[district_col].value) if district_col is not None and row[district_col].value else ''
            address = str(row[address_col].value) if address_col is not None and row[address_col].value else ''
            phone = str(row[phone_col].value) if phone_col is not None and row[phone_col].value else ''
            
            # Skip empty rows
            if not name or name == 'None':
                continue
            
            # Clean up name - add company prefix if not present
            if company.split()[0] not in name:
                name = f"{company} {name}"
            
            # Create Google Maps URL
            search_query = f"{name} {address} {city}"
            google_maps_url = f"https://www.google.com/maps/search/?api=1&query={search_query.replace(' ', '+')}"
            
            branch = {
                'id': str(uuid.uuid4()),
                'name': name.strip(),
                'company': company,
                'city': city.strip(),
                'district': district.strip(),
                'address': address.strip(),
                'phone': phone.strip(),
                'google_maps_url': google_maps_url,
                'logo_url': '',
                'working_hours': {},
                'source_url': '',
                'created_at': datetime.utcnow()
            }
            branches.append(branch)
            
        except Exception as e:
            print(f"Error processing row {row_idx}: {e}")
            continue
    
    return branches

async def import_branches():
    """Main import function"""
    total_imported = 0
    
    for company, url in EXCEL_FILES.items():
        print(f"\n{'='*50}")
        print(f"Processing {company}...")
        print(f"{'='*50}")
        
        try:
            # Download file
            filename = f"{company.replace(' ', '_')}.xlsx"
            filepath = await download_file(url, filename)
            print(f"Downloaded to {filepath}")
            
            # Extract branches
            branches = extract_branches_from_excel(filepath, company)
            print(f"Extracted {len(branches)} branches")
            
            if branches:
                # First delete existing branches for this company
                delete_result = await db.branches.delete_many({'company': company})
                print(f"Deleted {delete_result.deleted_count} existing {company} branches")
                
                # Insert new branches
                result = await db.branches.insert_many(branches)
                print(f"Inserted {len(result.inserted_ids)} branches for {company}")
                total_imported += len(result.inserted_ids)
            
            # Cleanup
            os.remove(filepath)
            
        except Exception as e:
            print(f"Error processing {company}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print(f"\n{'='*50}")
    print(f"TOTAL IMPORTED: {total_imported} branches")
    print(f"{'='*50}")
    
    # Get total count in database
    total_count = await db.branches.count_documents({})
    print(f"Total branches in database: {total_count}")

if __name__ == '__main__':
    asyncio.run(import_branches())
